import system_configuration as conf
# Get the parameters either from main parameters.py, or from a previous simulation
if(conf.RUN_SIMULATION is True):
    import parameters as env
else:
    import importlib
    env = importlib.import_module('parameters', conf.RUN_SIMULATION)

import learning as l
import energy as e
import time
import pandas as pd
import numpy as np
import openpyxl
from shutil import copyfile
from os import path
from itertools import product


def getAllSimulationPossibilities():
    global TOTAL_SIMULATIONS
    global COMMON_N_PATTERNS_X_PATTERN_FEATURES

    # Only take N_PATTERNS found in X_PATTERN_FEATURES, if instructed.
    if(env.ENSURE_N_PATTERNS_EQUALS_X_PATTERNS_FEATURES):
        COMMON_N_PATTERNS_X_PATTERN_FEATURES = list(
            set(env.N_PATTERNS) & set(env.X_PATTERN_FEATURES))
        env.N_PATTERNS = COMMON_N_PATTERNS_X_PATTERN_FEATURES
        env.X_PATTERN_FEATURES = COMMON_N_PATTERNS_X_PATTERN_FEATURES
    else:
        COMMON_N_PATTERNS_X_PATTERN_FEATURES = []

    # Returns a list of all possible permutations of the parameters for looping later.
    allSimulations = list(product(env.CACHE_ALGORITHMS, env.X_PATTERN_FEATURES, env.N_PATTERNS,
                                  env.LEARNING_RATES, env.MAX_SIZES_OF_TRANSIENT_MEMORY, env.MAINTENANCE_COSTS_OF_TRANSIENT_MEMORY,
                                  env.DECAY_TAUS_OF_TRANSIENT_MEMORY))
    TOTAL_SIMULATIONS = len(allSimulations) * len(env.SEEDS)
    return allSimulations


def simulate(simulationNumber, simulationTypeNumber, totalSimulations, cacheAlgorithm, xPatternFeature, nPattern, learningRate, maxSizeOfTransientMemory, maintenanceCostOfTransientMemory, decayTauOfTransientMemory, seed, filePath, directoryName):

    env.setCacheAlgorithm(cacheAlgorithm)
    env.setXPatternFeature(xPatternFeature)
    env.setNPattern(nPattern)
    env.setLearningRate(learningRate)
    env.setMaintenaceCostOfTransientMemory(maintenanceCostOfTransientMemory)
    env.setDecayTauOfTransientMemory(decayTauOfTransientMemory)
    env.setMaxSizeOfTransientMemory(maxSizeOfTransientMemory)
    env.setSeed(seed)
    env.setWeightModel()  # must always be the last thing to be called!

    #TODO: Refactor this so it doesnt have to rerun. If max_sizes is empty, find the optimal.
    if(env.MAX_SIZES_OF_TRANSIENT_MEMORY[0] == 'optimal'):
        env.MAX_SIZE_OF_TRANSIENT_MEMORY = e.calculateOptimalThreshold()
        env.setMaxSizeOfTransientMemory(env.MAX_SIZE_OF_TRANSIENT_MEMORY)
        env.setWeightModel()

    start = time.time()

    # Generate datasets
    trainingDatasetX, trainingDatasetY, testingDatasetX, testingDatasetY = l.getDatasets()

    # Use training dataset to estimate weights, where each row of weights is a new epoch.
    epochIndexForConvergence, weightsByEpoch, consolidationsByEpoch = l.trainWeights(
        trainingDatasetX, trainingDatasetY)

    # Get performance metrics for seen data
    trainNLL, trainAccuracy, trainPrecision, trainSensitivity, trainSpecificity = l.testWeights(
        trainingDatasetX, trainingDatasetY, weightsByEpoch)
    # Get performance metrics for UNSEEN data
    testNLL, testAccuracy, testPrecision, testSensitivity, testSpecificity = l.testWeights(
        testingDatasetX, testingDatasetY, weightsByEpoch)

    # Calculate energy expenditure
    theoreticalEfficiency = e.calculateTheoreticalEfficiency()
    metabolicEnergy = e.calculateMetabolicEnergy(weightsByEpoch)
    theoreticalMinimumEnergy = e.calculateTheoreticalMinimumEnergy(
        weightsByEpoch)
    efficiency = e.calculateEfficiency(
        metabolicEnergy, theoreticalMinimumEnergy)
    maintenanceEnergy = e.calculateEnergyFromMaintenance(
        weightsByEpoch)
    consolidationEnergy = e.calculateEnergyFromConsolidations(
        consolidationsByEpoch)
    energyBeforeThr = e.calculateEnergyJustBeforeThreshold(weightsByEpoch, consolidationsByEpoch)
    optimalThreshold = e.calculateOptimalThreshold()
    report = {
        simulationNumber:
        {
            'time_elapsed': str(time.time() - start),
            'simulationTypeNumber': simulationTypeNumber,
            'seed': seed,
            'learning_rate': learningRate,
            'n_pattern': nPattern,
            'n_pattern_features': xPatternFeature,  # TODO: notice it is x not n
            'max_epochs': env.MAX_EPOCHS,
            'energy_exponent': env.ENERGY_EXPONENT,
            'preset_simulation': env.PRESET_SIMULATION,
            'Learning was complete at epoch #': epochIndexForConvergence,
            'Theoretical: random-walk efficiency': str(theoreticalEfficiency),
            'Simulated: efficiency (m_perc/m_min)': str(efficiency),
            'Simulated-Theoretical efficiency difference': str(round((efficiency - theoreticalEfficiency), 3)),
            'Theoretical: minimum energy for learning': str(theoreticalMinimumEnergy),
            # Energy used hypothetically with no caching
            'Simulated: energy actually used by learning': str(metabolicEnergy),
            'Energy expended by simulations for consolidations': str(consolidationEnergy),
            'Energy expended by simulations for maintenance': str(maintenanceEnergy),
            'Energy expended by simulations for maintenance (plus before thr)': str(maintenanceEnergy + energyBeforeThr),
            'Energy expended total': str(round((consolidationEnergy + maintenanceEnergy), 3)),
            '(seen) NLL': str(trainNLL)+'%',
            '(seen) Accuracy': str(trainAccuracy)+'%',
            '(seen) Precision': str(trainPrecision)+'%',
            '(seen) Sensitivity': str(trainSensitivity)+'%',
            '(seen) Specificity': str(trainSpecificity)+'%',
            '(unseen) NLL': str(testNLL)+'%',
            '(unseen) Accuracy': str(testAccuracy)+'%',
            '(unseen) Precision': str(testPrecision)+'%',
            '(unseen) Sensitivity': str(testSensitivity)+'%',
            '(unseen) Specificity': str(testSpecificity)+'%',
            'weights_initialised_as': str(env.WEIGHTS_INITIALISED_AS),
            'cache_algorithm': str(env.CACHE_ALGORITHM),
            'max_size_of_transient_memory': str(maxSizeOfTransientMemory),
            'maintenance_cost_of_transient_memory': str(maintenanceCostOfTransientMemory),
            'Optimal threshold': str(optimalThreshold),
            'Decay rate of transient memory': str(decayTauOfTransientMemory),
            'number of consolidations': str(len(np.nonzero(consolidationsByEpoch))),
            'mean consolidation size': str(consolidationsByEpoch[np.nonzero(consolidationsByEpoch)].mean()),
        }
    }

    # Copy the parameters file too (so output can be reproduced)
    # TODO: we lazily remove output.csv to get the folder
    copyfile('parameters.py', str(filePath[:-10])+'parameters.py')
    copyfile('__init__.py', str(filePath[:-10])+'__init__.py')

    if path.isfile(filePath):
        mode = 'a'
        header = 0
        columns = list(report[simulationNumber].keys())
    else:
        mode = 'w'
        header = list(report[simulationNumber].keys())
        columns = header
    pd.DataFrame.from_dict(report).transpose().to_csv(
        filePath, header=header, columns=columns, mode=mode)

    if((env.STORE_WEIGHTS_TO_SPREADSHEET == False) or (totalSimulations > 50)):
        pass
    else:
        # Reorder weights, so that now they are by weight memory type.
        # eg. [0] =  weight1, weight2, ..., weightK
        # Where weight1 = t1, t2, ..., tn
        weightsByEpoch = np.swapaxes(weightsByEpoch, 0, 2)

        for weightMemoryTypeId, weightMemoryType in enumerate(weightsByEpoch):
            fileName = directoryName+'/data/simulation-'+str(simulationTypeNumber)+'-weights-' + \
                str(env.SYNAPSE_MEMORY_TYPES[weightMemoryTypeId]
                    ['name'])+'.xlsx'

            if path.exists(fileName):
                mode = 'a'
            else:
                wb = openpyxl.Workbook()
                ws = wb.worksheets[0]
                ws['A4'] = "The following sheets show each weight over time. The rows represent each weight, and the columns are epochs. Note that some seeds may have fewer epochs than others."
                ws['C1'] = 'simulationTypeNumber'
                ws['D1'] = 'seed'
                ws['E1'] = 'learning_rate'
                ws['F1'] = 'n_pattern'
                ws['G1'] = 'n_pattern_features'
                ws['H1'] = 'max_epochs'
                ws['I1'] = 'energy_exponent'
                ws['J1'] = 'preset_simulation'
                ws['K1'] = 'weights_initialised_as'
                ws['L1'] = 'cache_algorithm'
                ws['M1'] = 'max_size_of_transient_memory'
                ws['N1'] = 'maintenance_cost_of_transient_memory'
                ws['O1'] = 'Decay rate of transient memory'
                ws['C2'] = simulationTypeNumber
                ws['D2'] = seed
                ws['E2'] = learningRate
                ws['F2'] = nPattern
                ws['G2'] = xPatternFeature
                ws['H2'] = env.MAX_EPOCHS
                ws['I2'] = env.ENERGY_EXPONENT
                ws['J2'] = env.PRESET_SIMULATION
                ws['K2'] = env.WEIGHTS_INITIALISED_AS
                ws['L2'] = env.CACHE_ALGORITHM
                ws['M2'] = env.MAX_SIZE_OF_TRANSIENT_MEMORY
                ws['N2'] = env.MAINTENANCE_COST_OF_TRANSIENT_MEMORY
                ws['O2'] = env.DECAY_TAU_OF_TRANSIENT_MEMORY

                wb.save(fileName)
                wb.close()
                mode = 'a'

            excel_book = openpyxl.load_workbook(fileName)
            with pd.ExcelWriter(fileName, mode=mode) as writer:
                writer.book = excel_book
                writer.sheets = {
                    worksheet.title: worksheet for worksheet in excel_book.worksheets}
                pd.DataFrame(weightMemoryType[:, :epochIndexForConvergence]).dropna().to_excel(
                    writer, sheet_name='Seed-'+str(seed))

    print("Simulation "+str(simulationNumber) +
          " of "+str(totalSimulations))
    timeElapsed = time.time() - start
    print("Finished simulation (time elapsed: "+str(timeElapsed))
